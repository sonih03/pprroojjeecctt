import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_excel_report(title: str, n8n_data: dict, output_dir: str = "data") -> str:
    os.makedirs(output_dir, exist_ok=True)

    summary = n8n_data.get("summary", {}) if isinstance(n8n_data, dict) else {}
    rows = n8n_data.get("rows", []) if isinstance(n8n_data, dict) else (n8n_data if isinstance(n8n_data, list) else [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive_Summary"

    # 눈금선 표시
    ws.views.sheetView[0].showGridLines = True

    # -------------------------------------------------------------
    # 1. 메인 타이틀 영역
    # -------------------------------------------------------------
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 {title} - AI Executive Intelligence Report"
    title_cell.font = Font(name="맑은 고딕", size=16, bold=True, color="0F172A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # -------------------------------------------------------------
    # 2. KPI 대시보드 카드 영역 (3~5행)
    # -------------------------------------------------------------
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    card_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Card 1: 총 분석 이슈
    ws.merge_cells("A3:B3")
    ws.merge_cells("A4:B4")
    ws["A3"] = "분석 이슈 건수"
    ws["A4"] = f"{len(rows)} 건"

    # Card 2: HIGH 위험 건수
    high_count = sum(1 for r in rows if str(r.get("importance")).upper() == "HIGH")
    ws.merge_cells("C3:D3")
    ws.merge_cells("C4:D4")
    ws["C3"] = "주요 위험 (HIGH)"
    ws["C4"] = f"{high_count} 건"

    # Card 3: 시장 스탠스
    ws.merge_cells("E3:F3")
    ws.merge_cells("E4:F4")
    ws["E3"] = "시장 종합 스탠스"
    ws["E4"] = summary.get("market_stance", "중립 / 관망")

    # KPI 카드 스타일 적용
    for col in ["A", "C", "E"]:
        ws[f"{col}3"].font = Font(name="맑은 고딕", size=9, bold=True, color="64748B")
        ws[f"{col}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}4"].font = Font(name="맑은 고딕", size=14, bold=True, color="0F172A")
        ws[f"{col}4"].alignment = Alignment(horizontal="center", vertical="center")

    ws["C4"].font = Font(name="맑은 고딕", size=14, bold=True, color="DC2626")  # HIGH 건수는 빨간색

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 28

    # 경영진 총평 (Executive Takeaway) Box (6행)
    ws.merge_cells("A6:G6")
    takeaway = summary.get("executive_takeaway", "실시간 수집된 뉴스를 기반으로 기술 및 시장 영향력을 분석한 요약 리포트입니다.")
    ws["A6"] = f"💡 경영진 핵심 한줄 총평: {takeaway}"
    ws["A6"].font = Font(name="맑은 고딕", size=10, italic=True, bold=True, color="1E3A8A")
    ws["A6"].fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    ws["A6"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A6"].border = card_border
    ws.row_dimensions[6].height = 30

    # -------------------------------------------------------------
    # 3. 상세 분석 데이터 표 (8행부터)
    # -------------------------------------------------------------
    headers = ["번호", "중요도", "대응 긴급도", "카테고리", "주요 이슈 및 핵심 지표", "배경 및 현황", "산업 영향 및 추천 대응 전략"]

    ws.append([])  # 7행 빈 줄
    ws.append(headers)  # 8행 헤더
    ws.row_dimensions[8].height = 28

    # 헤더 스타일
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    for cell in ws[8]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 작성 (9행부터)
    for idx, item in enumerate(rows, 1):
        importance = str(item.get("importance", "MEDIUM")).upper()
        urgency = str(item.get("urgency", "단기 관찰"))
        category = str(item.get("category", "일반"))
        topic = str(item.get("topic", "-"))
        metric = item.get("key_metric")

        topic_text = f"{topic}\n(📊 핵심지표: {metric})" if metric else topic
        background = str(item.get("background") or item.get("content") or "-")

        impact = item.get("impact", "-")
        action = item.get("action_item", "-")
        detail_text = f"【시장/기술 영향】\n{impact}\n\n【추천 대응 전략】\n{action}"

        ws.append([idx, importance, urgency, category, topic_text, background, detail_text])

    # -------------------------------------------------------------
    # 4. 데이터 세부 스타일 및 색상 배지(Badge) 처리
    # -------------------------------------------------------------
    regular_font = Font(name="맑은 고딕", size=9.5)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # 파스텔 톤 배지 Fill 설정
    fill_high = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # 연분홍
    fill_med = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # 연노랑
    fill_low = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # 연초록

    font_high = Font(name="맑은 고딕", size=9.5, bold=True, color="991B1B")
    font_med = Font(name="맑은 고딕", size=9.5, bold=True, color="92400E")
    font_low = Font(name="맑은 고딕", size=9.5, bold=True, color="166534")

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=7):
        ws.row_dimensions[row[0].row].height = 85  # 가독성을 위한 셀 높이

        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if cell.column in [1, 2, 3, 4] else "left",
                vertical="top",
                wrap_text=True
            )

            # 중요도(B열) 색상 배지 처리
            if cell.column == 2:
                if cell.value == "HIGH":
                    cell.fill = fill_high
                    cell.font = font_high
                elif cell.value == "MEDIUM":
                    cell.fill = fill_med
                    cell.font = font_med
                else:
                    cell.fill = fill_low
                    cell.font = font_low

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 7  # 번호
    ws.column_dimensions['B'].width = 12  # 중요도
    ws.column_dimensions['C'].width = 13  # 긴급도
    ws.column_dimensions['D'].width = 16  # 카테고리
    ws.column_dimensions['E'].width = 32  # 주요 이슈 및 지표
    ws.column_dimensions['F'].width = 42  # 배경
    ws.column_dimensions['G'].width = 48  # 영향 및 전략

    file_path = os.path.join(output_dir, f"{title}_리포트.xlsx")
    wb.save(file_path)
    return file_path